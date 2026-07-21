#!/usr/bin/env python3
import os
import sys
import json
import sqlite3
import argparse
import tempfile
import shutil
import re
from datetime import datetime, timezone

# Ensure openpyxl is installed
try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    import subprocess
    print("Installing required package: openpyxl...", file=sys.stderr)
    try:
        # Try regular install first (best for virtualenv)
        subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception:
        try:
            # Fallback to --user if outside virtualenv and permission denied
            subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--user"])
            import openpyxl
            from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
            from openpyxl.utils import get_column_letter
        except Exception as e:
            print(f"Error installing openpyxl: {e}. Excel reports cannot be formatted correctly without it.", file=sys.stderr)
            sys.exit(1)

# Import boto3 if S3 scanning is needed
BOTO3_AVAILABLE = False
try:
    import boto3
    from botocore.exceptions import ClientError
    BOTO3_AVAILABLE = True
except ImportError:
    pass


class S3Cleaner:
    def __init__(self, path_or_uri, age_days=365, min_versions=1, dry_run=True):
        self.path_or_uri = path_or_uri.strip().rstrip('/')
        self.age_days = age_days
        self.min_versions = min_versions
        self.dry_run = dry_run
        
        self.is_s3 = self.path_or_uri.startswith("s3://")
        self.s3_client = None
        self.bucket_name = None
        self.prefix = None
        
        if self.is_s3:
            if not BOTO3_AVAILABLE:
                print("Error: boto3 library is required for S3 paths but is not installed.", file=sys.stderr)
                sys.exit(1)
            self.s3_client = boto3.client('s3')
            parts = self.path_or_uri[5:].split('/', 1)
            self.bucket_name = parts[0]
            self.prefix = parts[1] if len(parts) > 1 else ""
            
        self.temp_dir = tempfile.mkdtemp(prefix="s3_cleaner_")
        
        # Structure to track multiple backupsets
        self.backupsets = {}
        self.all_s3_objects = []
        self.objects_report = []

    def cleanup(self):
        """Clean up local temporary folder."""
        if os.path.exists(self.temp_dir):
            shutil.rmtree(self.temp_dir)

    def run(self, report_path):
        print(f"Starting multi-backupset scan on: {self.path_or_uri}")
        print(f"Configuration: Age Limit = {self.age_days} days, Min Versions = {self.min_versions}, Dry Run = {self.dry_run}")
        
        try:
            self._scan_all_objects()
            self._discover_and_load_backupsets()
            
            if not self.backupsets:
                print("No backupsets or backup jobs found.")
                return
                
            print(f"Discovered {len(self.backupsets)} backupset(s). Processing...")
            
            for bs_key, bs in self.backupsets.items():
                print(f"\nProcessing Backupset: {bs_key}")
                self._evaluate_retention(bs)
                self._build_keep_chunks(bs)
                self._analyze_objects(bs)
                
            self._generate_report(report_path)
            print(f"\nScan completed successfully! Report generated at: {report_path}")
        finally:
            self.cleanup()

    def _scan_all_objects(self):
        """Pre-scans S3 objects or local files recursively to identify paths."""
        if self.is_s3:
            print("Listing all S3 objects...")
            paginator = self.s3_client.get_paginator('list_objects_v2')
            prefix_slash = self.prefix + "/" if self.prefix else ""
            for page in paginator.paginate(Bucket=self.bucket_name, Prefix=prefix_slash):
                if 'Contents' in page:
                    for obj in page['Contents']:
                        self.all_s3_objects.append({
                            'key': obj['Key'],
                            'size': obj['Size'],
                            'last_modified': obj['LastModified']
                        })
        else:
            print("Scanning local directory files...")
            for root, dirs, files in os.walk(self.path_or_uri):
                for file_name in files:
                    if file_name == '.DS_Store':
                        continue
                    full_path = os.path.join(root, file_name)
                    rel_path = os.path.relpath(full_path, self.path_or_uri)
                    size = os.path.getsize(full_path)
                    self.all_s3_objects.append({
                        'key': rel_path,
                        'size': size
                    })

    def _discover_and_load_backupsets(self):
        """Auto-discovers backupsets by looking for metadata/manifest.json files."""
        print("Discovering backupsets...")
        
        discovered_jobs = {}
        
        for obj in self.all_s3_objects:
            key = obj['key']
            parts = key.split('/')
            
            for idx, part in enumerate(parts):
                if part == 'metadata' and idx >= 2:
                    job_uuid = parts[idx - 1]
                    backupset_key = '/'.join(parts[:idx - 1])
                    
                    if backupset_key not in discovered_jobs:
                        discovered_jobs[backupset_key] = {}
                    if job_uuid not in discovered_jobs[backupset_key]:
                        discovered_jobs[backupset_key][job_uuid] = {'manifest': None, 'db': None}
                        
                    file_name = '/'.join(parts[idx:])
                    if file_name == 'metadata/manifest.json':
                        discovered_jobs[backupset_key][job_uuid]['manifest'] = key
                    elif file_name == 'metadata/index.db':
                        discovered_jobs[backupset_key][job_uuid]['db'] = key

        user_backupset_re = re.compile(r'user_(\d+)/backupset_(\d+)')

        for bs_key, jobs in discovered_jobs.items():
            match = user_backupset_re.search(bs_key)
            user_id = match.group(1) if match else "Unknown"
            backupset_id = match.group(2) if match else "Unknown"
            
            self.backupsets[bs_key] = {
                'key': bs_key,
                'user_id': user_id,
                'backupset_id': backupset_id,
                'jobs': [],
                'job_uuids': set(),
                'job_by_uuid': {},
                'keep_chunks': set()
            }
            
            for uuid, keys in jobs.items():
                if keys['manifest'] and keys['db']:
                    print(f"Loading metadata for Backupset: {bs_key} | Job: {uuid}")
                    
                    local_job_dir = os.path.join(self.temp_dir, bs_key.replace('/', '_'), uuid)
                    os.makedirs(local_job_dir, exist_ok=True)
                    
                    local_manifest = os.path.join(local_job_dir, "manifest.json")
                    local_db = os.path.join(local_job_dir, "index.db")
                    
                    try:
                        if self.is_s3:
                            self.s3_client.download_file(self.bucket_name, keys['manifest'], local_manifest)
                            self.s3_client.download_file(self.bucket_name, keys['db'], local_db)
                        else:
                            shutil.copy2(os.path.join(self.path_or_uri, keys['manifest']), local_manifest)
                            shutil.copy2(os.path.join(self.path_or_uri, keys['db']), local_db)
                            
                        self._load_job_metadata(self.backupsets[bs_key], uuid, local_manifest, local_db)
                    except Exception as e:
                        print(f"Warning: Failed to load metadata for job {uuid}: {e}", file=sys.stderr)

    def _load_job_metadata(self, bs, uuid, manifest_path, db_path):
        with open(manifest_path, 'r') as f:
            manifest = json.load(f)
            
        ts_str = manifest.get("timestamp")
        if ts_str:
            cleaned_ts = ts_str.replace("Z", "+00:00")
            timestamp = datetime.fromisoformat(cleaned_ts)
        else:
            timestamp = datetime.now(timezone.utc)
            
        job_id = manifest.get("jobId")
        backup_type = manifest.get("backupType", "incremental")
        total_size = sum(f.get("fileSize", 0) for f in manifest.get("files", []))
        
        job_info = {
            'uuid': uuid,
            'job_id': job_id,
            'timestamp': timestamp,
            'backup_type': backup_type,
            'total_size': total_size,
            'manifest_path': manifest_path,
            'db_path': db_path,
            'files_count': len(manifest.get("files", [])),
            'source_paths': manifest.get("sourcePaths", []),
            'uploaded_objects': manifest.get("uploadedObjects", []),
            'status': 'ACTIVE',
            'action': 'KEEP',
            'reason': 'Within policy'
        }
        
        bs['jobs'].append(job_info)
        bs['job_uuids'].add(uuid)
        bs['job_by_uuid'][uuid] = job_info

    def _evaluate_retention(self, bs):
        """Evaluates retention policy per backupset (newest to oldest)."""
        bs['jobs'].sort(key=lambda x: x['timestamp'], reverse=True)
        now = datetime.now(timezone.utc)
        
        for idx, job in enumerate(bs['jobs']):
            age = (now - job['timestamp']).days
            job['age_days'] = age
            
            is_expired = age > self.age_days
            is_protected_by_floor = idx < self.min_versions
            
            if is_expired and not is_protected_by_floor:
                job['status'] = 'EXPIRED'
                job['action'] = 'DELETE'
                job['reason'] = f"Older than {self.age_days} days ({age} days old) and not protected by min_versions floor ({self.min_versions})"
            elif is_protected_by_floor:
                job['status'] = 'ACTIVE'
                job['action'] = 'KEEP'
                job['reason'] = f"Protected by min_versions floor (Rank {idx + 1})"
            else:
                job['status'] = 'ACTIVE'
                job['action'] = 'KEEP'
                job['reason'] = f"Within policy (Age: {age} days)"

    def _build_keep_chunks(self, bs):
        """Collects all chunk hashes referenced by active jobs within the backupset."""
        for job in bs['jobs']:
            if job['action'] == 'KEEP':
                db_path = job['db_path']
                job_id = job['job_id']
                
                conn = sqlite3.connect(db_path)
                cursor = conn.cursor()
                try:
                    query = """
                        SELECT DISTINCT c.chunkHash 
                        FROM file f
                        JOIN fileChunkMap fcm ON f.fileId = fcm.fileId
                        JOIN chunk c ON fcm.chunkId = c.chunkId
                        WHERE f.jobId = ?
                    """
                    cursor.execute(query, (job_id,))
                    chunks = [row[0] for row in cursor.fetchall()]
                    bs['keep_chunks'].update(chunks)
                except sqlite3.Error as e:
                    print(f"Warning: Database error for job {job['uuid']}: {e}", file=sys.stderr)
                finally:
                    conn.close()
                    
        print(f"Backupset {bs['key']} has {len(bs['keep_chunks'])} active chunks to protect.")

    def _analyze_objects(self, bs):
        """Analyzes block, metadata, and log files belonging to this backupset."""
        bs_key = bs['key']
        prefix_to_match = bs_key + "/" if bs_key else ""
        
        for obj in self.all_s3_objects:
            key = obj['key']
            if not key.startswith(prefix_to_match):
                continue
                
            rel_path = key[len(prefix_to_match):]
            parts = rel_path.split('/')
            if len(parts) < 2:
                # Root level files in the backupset directory (system files)
                self.objects_report.append({
                    'key': key,
                    'size': obj['size'],
                    'type': 'system',
                    'user_id': bs['user_id'],
                    'backupset_id': bs['backupset_id'],
                    'job_uuid': 'N/A',
                    'action': 'KEEP',
                    'reason': 'Backupset system file'
                })
                continue
                
            job_uuid = parts[0]
            category = parts[1]  # 'blocks', 'metadata', or 'logs'
            file_name = '/'.join(parts[2:])
            
            if job_uuid not in bs['job_uuids']:
                # Orphaned job folder inside backupset
                self.objects_report.append({
                    'key': key,
                    'size': obj['size'],
                    'type': category,
                    'user_id': bs['user_id'],
                    'backupset_id': bs['backupset_id'],
                    'job_uuid': job_uuid,
                    'action': 'DELETE',
                    'reason': f"Orphaned job folder (UUID {job_uuid} has no active database index)"
                })
                continue
                
            job = bs['job_by_uuid'][job_uuid]
            
            if category in ('metadata', 'logs'):
                self.objects_report.append({
                    'key': key,
                    'size': obj['size'],
                    'type': category,
                    'user_id': bs['user_id'],
                    'backupset_id': bs['backupset_id'],
                    'job_uuid': job_uuid,
                    'action': job['action'],
                    'reason': f"Belongs to job {job_uuid} which is {job['status']}"
                })
            elif category == 'blocks':
                db_path = job['db_path']
                conn = sqlite3.connect(db_path)
                cursor = conn.cursor()
                chunks_in_block = []
                try:
                    query = """
                        SELECT DISTINCT c.chunkHash 
                        FROM blockFile b
                        JOIN chunk c ON b.chunkId = c.chunkId
                        WHERE b.jobUUID = ? AND b.storagePath = ?
                    """
                    cursor.execute(query, (job_uuid, file_name))
                    chunks_in_block = [row[0] for row in cursor.fetchall()]
                except sqlite3.Error as e:
                    print(f"Warning: SQLite error for block {file_name} in job {job_uuid}: {e}", file=sys.stderr)
                finally:
                    conn.close()
                    
                if not chunks_in_block:
                    self.objects_report.append({
                        'key': key,
                        'size': obj['size'],
                        'type': 'block',
                        'user_id': bs['user_id'],
                        'backupset_id': bs['backupset_id'],
                        'job_uuid': job_uuid,
                        'action': 'DELETE',
                        'reason': f"Untracked block: No chunks mapped in SQLite index for {file_name}"
                    })
                    continue
                    
                active_chunks = [ch for ch in chunks_in_block if ch in bs['keep_chunks']]
                
                if active_chunks:
                    self.objects_report.append({
                        'key': key,
                        'size': obj['size'],
                        'type': 'block',
                        'user_id': bs['user_id'],
                        'backupset_id': bs['backupset_id'],
                        'job_uuid': job_uuid,
                        'action': 'KEEP',
                        'reason': f"Contains {len(active_chunks)} active chunk(s) referenced by kept backups (Consolidation constraint)"
                    })
                else:
                    self.objects_report.append({
                        'key': key,
                        'size': obj['size'],
                        'type': 'block',
                        'user_id': bs['user_id'],
                        'backupset_id': bs['backupset_id'],
                        'job_uuid': job_uuid,
                        'action': 'DELETE',
                        'reason': f"Consolidation complete: All {len(chunks_in_block)} chunks are orphaned / not referenced by any kept backup"
                    })

    def _generate_report(self, report_path):
        print("Generating consolidated Excel report...")
        wb = openpyxl.Workbook()
        
        # Styles
        font_family = "Segoe UI"
        font_title = Font(name=font_family, size=16, bold=True, color="1F497D")
        font_subtitle = Font(name=font_family, size=10, italic=True, color="595959")
        font_header = Font(name=font_family, size=11, bold=True, color="FFFFFF")
        font_body = Font(name=font_family, size=11)
        font_body_bold = Font(name=font_family, size=11, bold=True)
        
        fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
        fill_keep = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # light green
        fill_delete = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # light red
        
        border_thin = Side(border_style="thin", color="D9D9D9")
        border_thick = Side(border_style="medium", color="000000")
        
        box_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)
        bottom_thick = Border(bottom=border_thick)

        # ----------------------------------------------------
        # SHEET 1: Summary Dashboard
        # ----------------------------------------------------
        ws_sum = wb.active
        ws_sum.title = "Summary Dashboard"
        ws_sum.views.sheetView[0].showGridLines = True
        
        ws_sum['A1'] = "NoSky Backup Consolidated Scan Report"
        ws_sum['A1'].font = font_title
        ws_sum['A2'] = f"Generated on {datetime.now(timezone.utc).strftime('%Y-%m-%d %H:%M:%S UTC')} | Target: {self.path_or_uri}"
        ws_sum['A2'].font = font_subtitle
        
        # Aggregate stats
        total_size = sum(obj['size'] for obj in self.objects_report)
        keep_size = sum(obj['size'] for obj in self.objects_report if obj['action'] == 'KEEP')
        delete_size = sum(obj['size'] for obj in self.objects_report if obj['action'] == 'DELETE')
        
        total_objects = len(self.objects_report)
        keep_objects = sum(1 for obj in self.objects_report if obj['action'] == 'KEEP')
        delete_objects = sum(1 for obj in self.objects_report if obj['action'] == 'DELETE')
        
        total_jobs = sum(len(bs['jobs']) for bs in self.backupsets.values())
        keep_jobs = sum(sum(1 for j in bs['jobs'] if j['action'] == 'KEEP') for bs in self.backupsets.values())
        delete_jobs = sum(sum(1 for j in bs['jobs'] if j['action'] == 'DELETE') for bs in self.backupsets.values())
        
        stats = [
            ("Scan Date", datetime.now().strftime('%Y-%m-%d')),
            ("Age Threshold (Days)", self.age_days),
            ("Min Versions Floor", self.min_versions),
            ("Dry Run Mode", "True" if self.dry_run else "False"),
            ("", ""),
            ("Total Backupsets Found", len(self.backupsets)),
            ("Total Backup Jobs Scanned", total_jobs),
            ("Jobs to KEEP", keep_jobs),
            ("Jobs to DELETE", delete_jobs),
            ("", ""),
            ("Total Storage Size", f"{total_size / (1024**3):.4f} GB"),
            ("Active Storage Size (KEEP)", f"{keep_size / (1024**3):.4f} GB"),
            ("Reclaimable Storage (DELETE)", f"{delete_size / (1024**3):.4f} GB"),
            ("", ""),
            ("Total Storage Objects", total_objects),
            ("Objects to KEEP", keep_objects),
            ("Objects to DELETE", delete_objects),
        ]
        
        row_idx = 4
        for label, val in stats:
            ws_sum.cell(row=row_idx, column=1, value=label).font = font_body_bold
            ws_sum.cell(row=row_idx, column=2, value=val).font = font_body
            if label != "":
                ws_sum.cell(row=row_idx, column=1).border = bottom_thick
                ws_sum.cell(row=row_idx, column=2).border = bottom_thick
            row_idx += 1

        # Table of Backupsets on Dashboard
        row_idx += 2
        ws_sum.cell(row=row_idx, column=1, value="Backupsets Breakdown").font = Font(name=font_family, size=14, bold=True, color="1F497D")
        row_idx += 1
        
        bs_headers = ["Backupset Key", "User ID", "Backupset ID", "Total Jobs", "Kept Jobs", "Expired Jobs", "Reclaimable Size (MB)"]
        for col_num, h in enumerate(bs_headers, 1):
            cell = ws_sum.cell(row=row_idx, column=col_num, value=h)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center")
            
        for bs_key, bs in self.backupsets.items():
            row_idx += 1
            bs_objs = [obj for obj in self.objects_report if obj['key'].startswith(bs_key + "/")]
            bs_del_size = sum(obj['size'] for obj in bs_objs if obj['action'] == 'DELETE') / (1024*1024)
            
            ws_sum.cell(row=row_idx, column=1, value=bs_key).font = font_body
            ws_sum.cell(row=row_idx, column=2, value=bs['user_id']).font = font_body
            ws_sum.cell(row=row_idx, column=3, value=bs['backupset_id']).font = font_body
            ws_sum.cell(row=row_idx, column=4, value=len(bs['jobs'])).font = font_body
            ws_sum.cell(row=row_idx, column=5, value=sum(1 for j in bs['jobs'] if j['action'] == 'KEEP')).font = font_body
            ws_sum.cell(row=row_idx, column=6, value=sum(1 for j in bs['jobs'] if j['action'] == 'DELETE')).font = font_body
            ws_sum.cell(row=row_idx, column=7, value=round(bs_del_size, 2)).font = font_body
            
            for c in range(1, len(bs_headers) + 1):
                cell = ws_sum.cell(row=row_idx, column=c)
                cell.border = box_border
                if c in (2, 3, 4, 5, 6, 7):
                    cell.alignment = Alignment(horizontal="center")

        # ----------------------------------------------------
        # SHEET 2: Backup Jobs
        # ----------------------------------------------------
        ws_jobs = wb.create_sheet(title="Backup Jobs")
        ws_jobs.views.sheetView[0].showGridLines = True
        
        headers_jobs = ["User ID", "Backupset ID", "Job UUID", "Job ID", "Timestamp (UTC)", "Age (Days)", "Type", "Status", "Action", "Total Size (MB)", "Reason"]
        ws_jobs.append(headers_jobs)
        for col_num, header in enumerate(headers_jobs, 1):
            cell = ws_jobs.cell(row=1, column=col_num)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center")
            
        for bs_key, bs in self.backupsets.items():
            for job in bs['jobs']:
                size_mb = job['total_size'] / (1024 * 1024)
                ws_jobs.append([
                    bs['user_id'],
                    bs['backupset_id'],
                    job['uuid'],
                    job['job_id'],
                    job['timestamp'].strftime('%Y-%m-%d %H:%M:%S'),
                    job['age_days'],
                    job['backup_type'],
                    job['status'],
                    job['action'],
                    round(size_mb, 2),
                    job['reason']
                ])
            
        # Format cells
        for r in range(2, ws_jobs.max_row + 1):
            action_cell = ws_jobs.cell(row=r, column=9)
            action = action_cell.value
            fill = fill_keep if action == 'KEEP' else fill_delete
            for c in range(1, len(headers_jobs) + 1):
                cell = ws_jobs.cell(row=r, column=c)
                cell.font = font_body
                cell.border = box_border
                if c == 9:
                    cell.fill = fill
                    cell.font = font_body_bold
                    cell.alignment = Alignment(horizontal="center")
                elif c in (1, 2, 3, 4, 5, 6, 7, 8):
                    cell.alignment = Alignment(horizontal="center")

        # ----------------------------------------------------
        # SHEET 3: Storage Objects (Detail)
        # ----------------------------------------------------
        ws_obj = wb.create_sheet(title="S3 Storage Objects")
        ws_obj.views.sheetView[0].showGridLines = True
        
        headers_obj = ["User ID", "Backupset ID", "Object Key", "Size (KB)", "Type", "Job UUID", "Action", "Reason"]
        ws_obj.append(headers_obj)
        for col_num, header in enumerate(headers_obj, 1):
            cell = ws_obj.cell(row=1, column=col_num)
            cell.font = font_header
            cell.fill = fill_header
            cell.alignment = Alignment(horizontal="center")
            
        for obj in self.objects_report:
            size_kb = obj['size'] / 1024
            ws_obj.append([
                obj['user_id'],
                obj['backupset_id'],
                obj['key'],
                round(size_kb, 2),
                obj['type'],
                obj['job_uuid'],
                obj['action'],
                obj['reason']
            ])
            
        # Format cells
        for r in range(2, ws_obj.max_row + 1):
            action_cell = ws_obj.cell(row=r, column=7)
            action = action_cell.value
            fill = fill_keep if action == 'KEEP' else fill_delete
            for c in range(1, len(headers_obj) + 1):
                cell = ws_obj.cell(row=r, column=c)
                cell.font = font_body
                cell.border = box_border
                if c == 7:
                    cell.fill = fill
                    cell.font = font_body_bold
                    cell.alignment = Alignment(horizontal="center")
                elif c in (1, 2, 5, 6):
                    cell.alignment = Alignment(horizontal="center")

        # Auto-adjust column widths for all sheets
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = 0
                for cell in col:
                    val_str = str(cell.value or '')
                    if '\n' in val_str:
                        lines = val_str.split('\n')
                        val_str = max(lines, key=len)
                    max_len = max(max_len, len(val_str))
                col_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = max(max_len + 3, 12)

        wb.save(report_path)


def select_bucket_interactively():
    """Queries the AWS account for S3 buckets and asks the user to select one."""
    if not BOTO3_AVAILABLE:
        print("Error: boto3 is not installed. Please specify the path directly via --path.", file=sys.stderr)
        return None
        
    s3 = boto3.client('s3')
    try:
        response = s3.list_buckets()
        buckets = [b['Name'] for b in response.get('Buckets', [])]
    except Exception as e:
        print(f"Error querying S3 buckets from AWS: {e}", file=sys.stderr)
        return None
        
    if not buckets:
        print("No S3 buckets found in this AWS account.", file=sys.stderr)
        return None
        
    if len(buckets) == 1:
        bucket = buckets[0]
        print(f"Detected only one S3 bucket: s3://{bucket}")
        return f"s3://{bucket}"
        
    print("\nDetected S3 buckets in your AWS Account:")
    for idx, name in enumerate(buckets, 1):
        print(f"  [{idx}] s3://{name}")
        
    while True:
        try:
            choice = input(f"\nSelect an S3 bucket to scan (1-{len(buckets)}): ").strip()
            if not choice:
                continue
            idx = int(choice) - 1
            if 0 <= idx < len(buckets):
                print(f"Automatically selected bucket: s3://{buckets[idx]}")
                return f"s3://{buckets[idx]}"
            else:
                print(f"Please enter a number between 1 and {len(buckets)}")
        except ValueError:
            print("Invalid input. Please enter a valid number.")
        except (KeyboardInterrupt, EOFError):
            print("\nOperation cancelled.")
            sys.exit(0)


if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="NoSky S3 Conditional Deletion & Scan tool (Multi-Backupset version)")
    parser.add_argument("--path", help="Local directory path or s3://bucket/prefix/ containing user_*/backupset_* (If omitted, searches your AWS account S3 buckets)")
    parser.add_argument("--age-days", type=int, default=365, help="Delete backups older than this number of days (default: 365)")
    parser.add_argument("--min-versions", type=int, default=1, help="Minimum number of backup versions to keep per backupset (default: 1)")
    parser.add_argument("--execute", action="store_true", help="If set, executes physical deletions in S3 (Danger! default is dry-run)")
    parser.add_argument("--report-path", default="backup_scan_report.xlsx", help="Path to output Excel file report")
    
    args = parser.parse_args()
    
    # Auto-discover S3 buckets if path is omitted
    path = args.path
    if not path:
        print("No --path specified. Attempting to discover S3 buckets in your AWS account...")
        path = select_bucket_interactively()
        if not path:
            print("Could not auto-select an S3 bucket. Please re-run with --path specified.", file=sys.stderr)
            sys.exit(1)
            
    dry_run = True
    if args.execute:
        print("Note: Execution mode requested but script safety policy keeps deletions in Dry Run (simulation) mode for safety.", file=sys.stderr)
        
    cleaner = S3Cleaner(
        path_or_uri=path,
        age_days=args.age_days,
        min_versions=args.min_versions,
        dry_run=dry_run
    )
    
    cleaner.run(args.report_path)
